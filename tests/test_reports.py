from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from app.reports import ReportRow, build_report


def test_report_has_required_columns_and_marks() -> None:
    content = build_report(
        date(2026, 9, 1),
        [
            ReportRow("8МК", "Иванов Иван", True, False),
            ReportRow("2Б", "Петров Пётр", True, True),
        ],
    )
    workbook = load_workbook(BytesIO(content))
    sheet = workbook["Заказы"]

    assert [cell.value for cell in sheet[1]] == ["Класс", "ФИО", "Завтрак", "Обед"]
    assert list(sheet.values)[1:] == [
        ("2Б", "Петров Пётр", "✓", "✓"),
        ("8МК", "Иванов Иван", "✓", "—"),
    ]


def test_empty_report_is_still_valid() -> None:
    workbook = load_workbook(BytesIO(build_report(date(2026, 9, 1), [])))
    assert workbook["Заказы"].max_row == 2
