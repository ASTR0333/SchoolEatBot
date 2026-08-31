from datetime import date
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from app.config import Settings
from app.database import Database
from app.models import Parent
from app.scheduler import DailyScheduler
from app.service import BotService


class FakeMax:
    def __init__(self) -> None:
        self.messages: list[tuple[int, str, object]] = []
        self.callbacks: list[tuple[str, str]] = []
        self.excels: list[tuple[int, str, bytes, str]] = []

    async def send_message(self, user_id, text, buttons=None, **_kwargs) -> None:  # noqa: ANN001
        self.messages.append((user_id, text, buttons))

    async def answer_callback(self, callback_id: str, notification: str) -> None:
        self.callbacks.append((callback_id, notification))

    async def send_excel(self, user_id: int, filename: str, content: bytes, text: str) -> None:
        self.excels.append((user_id, filename, content, text))


def make_settings(**overrides) -> Settings:  # noqa: ANN003
    values = {
        "max_bot_token": "test-token",
        "creator_user_id": 100,
        "teacher_1_id": 200,
        "teacher_2_id": 300,
        "class_1": "8МК",
        "class_2": "2Б",
    }
    values.update(overrides)
    return Settings(**values)


async def make_service(
    database_path: Path, settings: Settings
) -> tuple[BotService, Database, FakeMax]:
    database = Database(f"sqlite+aiosqlite:///{database_path}")
    await database.create_schema()
    max_client = FakeMax()
    service = BotService(settings, database, max_client)  # type: ignore[arg-type]
    return service, database, max_client


def test_empty_teacher_ids_are_optional() -> None:
    settings = make_settings(teacher_1_id="", teacher_2_id="  ")

    assert settings.teacher_assignments == ()
    assert settings.report_recipients == ((100, None),)


async def test_scheduler_sends_creator_all_classes_and_teacher_own_class() -> None:
    class FakeReportService:
        settings = make_settings(teacher_2_id=None)

        def __init__(self) -> None:
            self.sent: list[tuple[int, str | None]] = []

        async def delivery_exists(self, _key: str) -> bool:
            return False

        async def send_report_to(
            self, user_id: int, _target: date, *, class_name: str | None = None
        ) -> None:
            self.sent.append((user_id, class_name))

        async def record_delivery(self, _key: str) -> None:
            pass

    service = FakeReportService()
    scheduler = DailyScheduler(service)  # type: ignore[arg-type]
    await scheduler._send_reports(date(2026, 9, 1))

    assert service.sent == [(100, None), (200, "8МК")]


async def test_creator_and_teacher_get_correct_report_scope(tmp_path: Path) -> None:
    settings = make_settings()
    service, database, max_client = await make_service(tmp_path / "bot.sqlite", settings)

    await service.upsert_parent({"user_id": 1, "first_name": "Первый"}, 1)
    await service.upsert_parent({"user_id": 2, "first_name": "Второй"}, 2)
    async with database.session() as session:
        first = await session.get(Parent, 1)
        second = await session.get(Parent, 2)
        assert first is not None and second is not None
        first.class_name = "8МК"
        first.child_name = "Иванов Иван"
        second.class_name = "2Б"
        second.child_name = "Петров Пётр"
        await session.commit()

    target = date(2026, 9, 1)
    await service.send_report_to(100, target)
    await service.send_report_to(200, target, class_name="8МК")

    creator_sheet = load_workbook(BytesIO(max_client.excels[0][2]))["Заказы"]
    teacher_sheet = load_workbook(BytesIO(max_client.excels[1][2]))["Заказы"]
    assert [row[1] for row in list(creator_sheet.values)[1:]] == ["Петров Пётр", "Иванов Иван"]
    assert [row[1] for row in list(teacher_sheet.values)[1:]] == ["Иванов Иван"]
    assert max_client.excels[1][0] == 200
    assert "8МК" in max_client.excels[1][1]

    await database.close()


async def test_staff_menu_does_not_require_child_profile(tmp_path: Path) -> None:
    settings = make_settings()
    service, database, max_client = await make_service(tmp_path / "bot.sqlite", settings)

    await service.upsert_parent({"user_id": 100, "first_name": "Создатель"}, 100)
    await service.upsert_parent({"user_id": 200, "first_name": "Учитель"}, 200)
    await service.send_menu(100)
    await service.send_menu(200)

    assert "всем классам" in max_client.messages[-2][1]
    assert "класса 8МК" in max_client.messages[-1][1]

    await database.close()
