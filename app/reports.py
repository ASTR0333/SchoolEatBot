from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


@dataclass(frozen=True)
class ReportRow:
    class_name: str
    child_name: str
    breakfast: bool
    lunch: bool


def build_report(target: date, rows: list[ReportRow]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Заказы"
    sheet.freeze_panes = "A2"
    sheet.append(["Класс", "ФИО", "Завтрак", "Обед"])

    for row in sorted(rows, key=lambda item: (item.class_name, item.child_name.casefold())):
        sheet.append(
            [
                row.class_name,
                row.child_name,
                "✓" if row.breakfast else "—",
                "✓" if row.lunch else "—",
            ]
        )

    if not rows:
        sheet.append(["—", "Нет зарегистрированных учеников", "—", "—"])

    header_fill = PatternFill("solid", fgColor="2F75B5")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in sheet.iter_rows(min_row=2, min_col=3, max_col=4):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")
            if cell.value == "✓":
                cell.font = Font(color="008000", bold=True, size=14)

    sheet.column_dimensions["A"].width = 13
    sheet.column_dimensions["B"].width = 38
    sheet.column_dimensions["C"].width = 14
    sheet.column_dimensions["D"].width = 14
    sheet.auto_filter.ref = f"A1:D{sheet.max_row}"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.oddHeader.center.text = f"Заказы на {target:%d.%m.%Y}"

    table = Table(displayName="Orders", ref=f"A1:D{sheet.max_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
